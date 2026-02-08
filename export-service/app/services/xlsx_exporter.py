"""XLSX exporter for audit logs."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Column headers for the audit log XLSX export.
AUDIT_COLUMNS = [
    "Timestamp",
    "User ID",
    "Action",
    "Resource Type",
    "Resource ID",
    "Workspace ID",
    "Details",
    "IP Address",
]

# Header styling
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="1A56DB", end_color="1A56DB", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


class XlsxExporter:
    """Exports audit log entries as an Excel (.xlsx) file."""

    def export_audit_logs(self, logs: list[dict]) -> BytesIO:
        """Generate an XLSX file from audit log records.

        Args:
            logs: List of audit log dicts, each containing keys matching
                  the column headers (snake_case).

        Returns:
            BytesIO buffer containing the generated XLSX file.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Logs"

        # ── Header row ────────────────────────────────────────────────
        for col_idx, header in enumerate(AUDIT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGNMENT

        # ── Data rows ─────────────────────────────────────────────────
        field_keys = [
            "timestamp",
            "user_id",
            "action",
            "resource_type",
            "resource_id",
            "workspace_id",
            "details",
            "ip_address",
        ]

        for row_idx, entry in enumerate(logs, start=2):
            for col_idx, key in enumerate(field_keys, start=1):
                value = entry.get(key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")

                # Apply date format to the timestamp column
                if key == "timestamp" and value:
                    cell.number_format = "YYYY-MM-DD HH:MM:SS"

        # ── Auto-adjust column widths ─────────────────────────────────
        for col_idx in range(1, len(AUDIT_COLUMNS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(AUDIT_COLUMNS[col_idx - 1])  # start with header width

            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

            # Add a small padding and cap at a reasonable width
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

        # ── Serialize to bytes ────────────────────────────────────────
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
