"""XLSX parser using openpyxl."""

import logging

from openpyxl import load_workbook

from app.parsers.base import BaseParser, ParsedSection

logger = logging.getLogger(__name__)


class XlsxParser(BaseParser):
    """Extracts structured sections from Excel XLSX workbooks.

    Each worksheet becomes a section group.  Rows are serialised to
    pipe-delimited text.  The first row of each sheet is assumed to be a
    header and is included in every section to provide column context.
    """

    async def parse(self, file_path: str) -> list[ParsedSection]:
        sections: list[ParsedSection] = []

        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        except Exception:
            logger.exception("Failed to open XLSX: %s", file_path)
            raise

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows: list[list[str]] = []

                for row in ws.iter_rows(values_only=True):
                    cell_values = [
                        str(cell).strip() if cell is not None else ""
                        for cell in row
                    ]
                    # Skip completely empty rows
                    if any(v for v in cell_values):
                        rows.append(cell_values)

                if not rows:
                    continue

                # Treat first row as header
                header = rows[0]
                header_text = " | ".join(header)

                # Group data rows into manageable sections (up to 50 rows per section)
                data_rows = rows[1:]
                if not data_rows:
                    # Only a header row - still emit it
                    sections.append(
                        ParsedSection(
                            content=header_text,
                            page_number=None,
                            section_title=sheet_name,
                            section_hierarchy=[sheet_name],
                            chunk_type="table",
                        )
                    )
                    continue

                batch_size = 50
                for batch_idx in range(0, len(data_rows), batch_size):
                    batch = data_rows[batch_idx : batch_idx + batch_size]
                    lines = [header_text]  # repeat header for context
                    for row_cells in batch:
                        lines.append(" | ".join(row_cells))

                    sections.append(
                        ParsedSection(
                            content="\n".join(lines),
                            page_number=None,
                            section_title=sheet_name,
                            section_hierarchy=[sheet_name],
                            chunk_type="table",
                        )
                    )
        finally:
            wb.close()

        logger.info("XLSX parsed: %d sections from %s", len(sections), file_path)
        return sections
